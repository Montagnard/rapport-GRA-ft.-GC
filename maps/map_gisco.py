# /// script
# requires-python = ">=3.11"
# dependencies = ["geopandas", "matplotlib", "numpy", "openpyxl", "shapely"]
# ///
"""
Carte EU-27 avec barres par pays — basemap GISCO Eurostat (local).

Usage:
    uv run work/maps/map_gisco.py
    uv run work/maps/map_gisco.py -o work/maps/map_gisco.png
    uv run work/maps/map_gisco.py --scale 10M -o work/maps/map_gisco.png
    uv run work/maps/map_gisco.py --scale 03M --dpi 300

Scales disponibles : 60M  20M  10M  03M  (fichiers dans work/maps/basemaps/)
"""

import argparse
from pathlib import Path

import geopandas as gpd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import Rectangle
from shapely.affinity import scale as geom_scale, translate as geom_translate
import openpyxl

# ── Palette ──────────────────────────────────────────────────────────────────
COLORS = {
    'FTT':         '#1F3864',
    'DST':       '#2E75B6',
    'AdTax':         '#70AD47',
    'Aviation Tax':      '#C55A11',
    'Wealth':'#7030A0',
    'Luxury VAT':  '#C00000',
}
LAND_COLOR   = '#F0EFE9'
BORDER_COLOR = '#BBBBBB'
OCEAN_COLOR  = None
#'#D6EAF8'

# ── Géométrie des barres (mètres, EPSG:3035) ─────────────────────────────────
BAR_W   = 50_000
BAR_GAP = 10_000
MAX_H   = 600_000

# ── EU-27 (ISO alpha-3) ───────────────────────────────────────────────────────
EU27 = {
    'AUT','BEL','BGR','HRV','CYP','CZE','DNK','EST','FIN','FRA',
    'DEU','GRC','HUN','IRL','ITA','LVA','LTU','LUX','MLT','NLD',
    'POL','PRT','ROU','SVK','SVN','ESP','SWE',
}

NAME_ISO = {
    'Austria':'AUT','Belgium':'BEL','Bulgaria':'BGR','Croatia':'HRV',
    'Cyprus':'CYP','Czechia':'CZE','Denmark':'DNK','Estonia':'EST',
    'Finland':'FIN','France':'FRA','Germany':'DEU','Greece':'GRC',
    'Hungary':'HUN','Ireland':'IRL','Italy':'ITA','Latvia':'LVA',
    'Lithuania':'LTU','Luxembourg':'LUX','Malta':'MLT','Netherlands':'NLD',
    'Poland':'POL','Portugal':'PRT','Romania':'ROU','Slovakia':'SVK',
    'Slovenia':'SVN','Spain':'ESP','Sweden':'SWE',
}

# ── Offsets manuels (mètres EPSG:3035) ───────────────────────────────────────
# Centroïdes GISCO (millions de m) :
#   BEL 3.939/3.074   NLD 4.014/3.252   LUX 4.040/2.970
#   DEU 4.340/3.108   CZE 4.705/2.977   AUT 4.626/2.732
#   SVK 5.018/2.892   SVN 4.702/2.570   HRV 4.839/2.456
#   DNK 4.314/3.662   EST 5.233/4.051   LVA 5.221/3.840
#   LTU 5.198/3.660   FIN 5.099/4.663   SWE 4.637/4.382
#   IRL 3.131/3.489   PRT 2.778/2.020
#   CYP/LUX/MLT → encarts (INSET_COUNTRIES)
OFFSETS = {
    'BEL': ( -90_000,     -50_000),
    'NLD': ( -140_000,    50_000),
    'DEU': (   -40_000,   40_000),
    'CZE': (   0_000,    40_000),
    'AUT': (-200_000,   -40_000),
    'POL': ( -80_000,   100_000),
    'SVK': ( 100_000,    80_000),
    'HUN': ( 40_000,  -20_000),
    'SVN': (   20_000,   -0_000),
    'HRV': ( 250_000,  -250_000),
    'ROU': (  20_000,   -50_000),
    'BGR': (   0_000,   -60_000),
    'ITA': (   0_000,  -100_000),
    'EST': ( 100_000,    30_000),
    'LVA': (  30_000,   0_000),
    'LTU': ( 0_000,  -30_000),
    'DNK': (   0_000,     0_000),
    'PRT': ( -50_000,     0_000),
    'ESP': (  50_000,     0_000),
    'SWE': (  -50_000,     -300_000),
    'FIN': (  50_000,     -80_000),
    'GRC': (   0_000,     -70_000),

}

# ── INSET_POS : position fixe des barres (FRA uniquement) ────────────────────
# FRA : geometry GISCO inclut les DOM, centroïde tiré vers l'Atlantique.
INSET_POS = {
    'FRA': (3_720_000, 2_510_000),
}

# ── Encarts mis à l'échelle (polygone déporté + redimensionné) ────────────────
# CYP, LUX, MLT : déportés dans le coin supérieur gauche, ordonnés alphabétiquement.
# Les polygones sont mis à la même hauteur (INSET_HEIGHT) puis placés côte à côte.
# Les barres sont centrées sur le centroïde du polygone résultant.
INSET_COUNTRIES = ['CYP', 'LUX', 'MLT']  # gauche → droite
INSET_HEIGHT    = 250_000   # ← paramètre : hauteur uniforme de chaque polygone (m)
INSET_CY        = 4_300_000 # y du centroïde de chaque encart (base des barres)
INSET_START_X   = 2_550_000 # x du bord gauche du premier encart
INSET_GAP       = 300_000   # espace horizontal entre encarts (m)

# ── Étendue de la carte (mètres EPSG:3035) ───────────────────────────────────
# Continentale : x 2.64M–5.95M, y 1.42M–5.31M (hors DOM et CYP déporté)
MAP_XLIM = (2_350_000, 6_050_000)  # ← rogner à droite : réduire x_max
MAP_YLIM = (1_300_000, 5_450_000)  # ← rogner en haut  : réduire y_max

# ── Labels %GNI au sommet des barres ─────────────────────────────────────────
BAR_LABEL_MIN_H  = 20_000  # hauteur minimale pour afficher le label (mètres)
BAR_LABEL_FSIZE  = 7       # taille de police des labels %GNI

# ── Coloration des pays par gain net ─────────────────────────────────────────
# None = couleur uniforme ; int = indice de colonne xlsx (ex: 19 pour le GNI)
NET_GAIN_COL  = None
NET_GAIN_CMAP = 'YlGn'


def place_insets(eu: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Scale INSET_COUNTRIES to INSET_HEIGHT and position them left-to-right."""
    eu = eu.copy()
    cur_x = INSET_START_X
    for iso in INSET_COUNTRIES:
        mask = eu['ADM0_A3'] == iso
        if not mask.any():
            continue
        idx  = eu.index[mask][0]
        geom = eu.at[idx, 'geometry']
        # Mise à l'échelle uniforme autour du centroïde
        b = geom.bounds          # (minx, miny, maxx, maxy)
        s = INSET_HEIGHT / (b[3] - b[1])
        cx0, cy0 = geom.centroid.x, geom.centroid.y
        geom = geom_scale(geom, s, s, origin=(cx0, cy0))
        # Positionnement : centroïde → (cur_x + w/2, INSET_CY)
        w  = geom.bounds[2] - geom.bounds[0]
        dx = (cur_x + w / 2) - geom.centroid.x
        dy = INSET_CY - geom.centroid.y
        eu.at[idx, 'geometry'] = geom_translate(geom, dx, dy)
        cur_x += w + INSET_GAP
    return eu


def load_world(scale: str = '20M') -> gpd.GeoDataFrame:
    basedir = Path(__file__).parent / 'basemaps'
    path = basedir / f'nuts0_{scale}_3035.geojson'
    if not path.exists():
        raise FileNotFoundError(
            f'Basemap introuvable : {path}\n'
            'Scales disponibles : ' + ', '.join(
                p.name.split('_')[1] for p in basedir.glob('nuts0_*_3035.geojson')
            )
        )
    gdf = gpd.read_file(path)
    gdf = gdf.rename(columns={'ISO3_CODE': 'ADM0_A3'})
    return gdf


def find_xlsx(script_dir: Path) -> Path:
    for candidate in [
        script_dir.parent / 'combined.xlsx',
        script_dir.parent.parent / 'work' / 'combined.xlsx',
    ]:
        if candidate.exists():
            return candidate
    raise FileNotFoundError('combined.xlsx non trouvé — passe le chemin en argument.')


def read_data(xlsx: Path) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    COL = {'FTT': 7, 'DST': 3,  'Aviation Tax': 11, 'AdTax': 5, 'Wealth': 9,  'Luxury VAT': 14}
    out = {}
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        if str(name).strip().lower() in ('european union', 'eu27', 'total'):
            continue
        iso = NAME_ISO.get(str(name).strip())
        if not iso:
            continue
        row = {
            t: float(v) if isinstance(v := ws.cell(r, c).value, (int, float)) else 0.0
            for t, c in COL.items()
        }
        row['_pct'] = {t: row[t] for t in COL}
        # Gain net (coloration choroplèthe)
        if NET_GAIN_COL is not None:
            net_v = ws.cell(r, NET_GAIN_COL).value
            row['_net'] = float(net_v) if isinstance(net_v, (int, float)) else None
        out[iso] = row
    return out


def draw_map(data: dict, world: gpd.GeoDataFrame, out: Path, dpi: int = 150):
    eu_real  = world[world['ADM0_A3'].isin(EU27)].copy()  # géométries originales
    eu_inset = place_insets(eu_real)                       # encarts scalés + déportés

    taxes      = list(COLORS)
    global_max = max(
        (data[iso][t] for iso in data for t in taxes if data[iso][t] > 0),
        default=1,
    )

    group_w = len(taxes) * BAR_W + (len(taxes) - 1) * BAR_GAP

    _bg = (0, 0, 0, 0) if OCEAN_COLOR is None else OCEAN_COLOR
    fig, ax = plt.subplots(figsize=(14, 11), facecolor=_bg)
    ax.set_facecolor(_bg)
    ax.set_aspect('equal')

    # Coloration choroplèthe (gain net)
    if NET_GAIN_COL is not None:
        net_vals = {iso: data[iso]['_net'] for iso in data
                    if data[iso].get('_net') is not None}
        vmin, vmax = min(net_vals.values()), max(net_vals.values())
        span = vmax - vmin or 1.0
        cmap = plt.get_cmap(NET_GAIN_CMAP)
        def _ccolor(iso):
            v = net_vals.get(iso)
            return cmap((v - vmin) / span) if v is not None else LAND_COLOR
        real_colors  = [_ccolor(r['ADM0_A3']) for _, r in eu_real.iterrows()]
        inset_colors = [_ccolor(r['ADM0_A3']) for _, r in eu_inset[eu_inset['ADM0_A3'].isin(INSET_COUNTRIES)].iterrows()]
    else:
        real_colors  = LAND_COLOR
        inset_colors = LAND_COLOR

    # Fond : vraies positions (LUX/MLT visibles à leur emplacement géographique)
    eu_real.plot(ax=ax, color=real_colors, edgecolor=BORDER_COLOR, linewidth=0.5, zorder=1)
    # Encarts : polygones scalés dans le coin haut-gauche, par-dessus le fond
    inset_mask = eu_inset['ADM0_A3'].isin(INSET_COUNTRIES)
    eu_inset[inset_mask].plot(ax=ax, color=inset_colors, edgecolor=BORDER_COLOR,
                              linewidth=0.5, zorder=2)

    # Centroïdes depuis les encarts (bonne position pour les barres)
    centroids = {r['ADM0_A3']: (r.geometry.centroid.x, r.geometry.centroid.y)
                 for _, r in eu_inset.iterrows()}

    for iso in data:
        if iso in INSET_POS:
            cx, cy = INSET_POS[iso]
        elif iso in centroids:
            dx, dy = OFFSETS.get(iso, (0, 0))
            cx = centroids[iso][0] + dx
            cy = centroids[iso][1] + dy
        else:
            continue

        x0 = cx - group_w / 2

        pct = data[iso].get('_pct', {})
        for i, tax in enumerate(taxes):
            val = data[iso][tax]
            if val <= 0:
                continue
            h = (val / global_max) * MAX_H
            x = x0 + i * (BAR_W + BAR_GAP)
            ax.add_patch(Rectangle(
                (x, cy), BAR_W, h,
                facecolor=COLORS[tax], edgecolor='white',
                linewidth=0.2, alpha=0.93, zorder=3,
            ))
            if h >= BAR_LABEL_MIN_H and tax in pct:
                ax.text(x + BAR_W / 2, cy + h + 4_000,
                        f'{pct[tax]:.2f}'.lstrip('0'),
                        fontsize=BAR_LABEL_FSIZE, rotation=90, ha='center', va='bottom',
                        color='#222222', zorder=6)

        ax.text(cx, cy - 22_000, iso,
                fontsize=10, ha='center', va='top', color='#333333',
                zorder=5, fontfamily='monospace')

    ax.set_xlim(*MAP_XLIM)
    ax.set_ylim(*MAP_YLIM)

    ax.set_xticks([])
    ax.set_yticks([])
    for sp in ax.spines.values():
        sp.set_visible(False)

    LABELS = {
        'DST':         'Digital Service Tax (3%)',
        'AdTax':       'Ad Tax (15%)',
        'FTT':         'Financial Transaction Tax (0.5%)',
        'Wealth':      'Wealth Tax',
        'Aviation Tax':'Aviation Tax',
        'Luxury VAT':  'Luxury VAT',
    }
    patches = [
        mpatches.Patch(facecolor=COLORS[t], edgecolor='white', label=LABELS[t])
        for t in taxes
    ]
    ax.legend(handles=patches, loc='lower left', frameon=True, framealpha=0.92,
              edgecolor='#CCCCCC', fontsize=8.5, title='Tax', title_fontsize=9)

    ax.text(0.99, 0.01,
            'Hauteur des barres : revenu / GNI pays, normalisé par taxe',
            transform=ax.transAxes, fontsize=6.5, color='#777777',
            ha='right', va='bottom')

    fig.tight_layout(pad=0.2)
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=dpi, bbox_inches='tight', transparent=(OCEAN_COLOR is None))
    print(f'Saved: {out}')
    plt.close(fig)


def main():
    script_dir = Path(__file__).parent
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', nargs='?', help='Chemin vers combined.xlsx (auto-détecté si absent)')
    ap.add_argument('-o', '--output', default=str(script_dir / 'map_gisco.png'))
    ap.add_argument('--dpi', type=int, default=150)
    ap.add_argument('--scale', default='20M',
                    choices=['60M', '20M', '10M', '03M'],
                    help='Résolution du basemap GISCO (défaut: 20M)')
    args = ap.parse_args()

    xlsx  = Path(args.xlsx) if args.xlsx else find_xlsx(script_dir)
    world = load_world(args.scale)
    draw_map(read_data(xlsx), world, Path(args.output), dpi=args.dpi)


if __name__ == '__main__':
    raise SystemExit(main())
