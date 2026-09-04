# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
Convert the combined xlsx into one LaTeX longtable per table-definition row.

Header section layout (rows 1–N, before "Austria"):
  Row 1 : source names (merged across column groups)
  Row 2 : column group names
  Row 3 : unit labels
  Row 4+ : table-definition rows — every row whose col A is non-empty becomes one
            output table; col A value = table name = output filename stem.
            Fill col B onward with a LaTeX label to include that column.
  First row whose col A = "Austria" → data section starts.

Usage:
    uv run work/xlsx2tex.py work/combined.xlsx -o work/report/table.tex
    uv run work/xlsx2tex.py work/combined.xlsx -o work/report/table.tex --caption "Tableau"
    uv run work/xlsx2tex.py work/combined.xlsx --stdout
"""

import argparse
import random
from collections import OrderedDict
from pathlib import Path

import openpyxl


ROW_SRC   = 1
ROW_COL   = 2
ROW_UNIT  = 3
ROW_LATEX = 4   # first possible table-definition row

TOTAL_KEYS = {'european union', 'eu27', 'total'}


def unescape_tex(s: str) -> str:
    return (s
        .replace(r'\%', '%').replace(r'\_', '_').replace(r'\&', '&')
        .replace(r'\#', '#').replace(r'\$', '$')
        .replace(r'\{', '{').replace(r'\}', '}')
    )


def escape_tex(s: str) -> str:
    s = unescape_tex(s)
    return (s
        .replace('&', r'\&')
        .replace('%', r'\%')
        .replace('_', r'\_')
        .replace('#', r'\#')
        .replace('$', r'\$')
        .replace('{', r'\{')
        .replace('}', r'\}')
        .replace('~', r'\textasciitilde{}')
        .replace('^', r'\textasciicircum{}')
    )


def fmt_num(val) -> str:
    if val is None or val == '':
        return ''
    if isinstance(val, float):
        if val == int(val):
            return f'{int(val):,}'
        return f'{val:,.1f}'
    if isinstance(val, int):
        return f'{val:,}'
    return escape_tex(str(val))


def propagate(row_values: list) -> list:
    """Forward-fill None values (handles merged cells in openpyxl reads)."""
    result, current = [], None
    for v in row_values:
        if v is not None:
            current = v
        result.append(current)
    return result


def read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    max_col = ws.max_column

    # Fixed metadata rows
    src_row = propagate([ws.cell(ROW_SRC, c).value for c in range(1, max_col + 1)])
    sources_by_col = {c: src_row[c - 1] for c in range(2, max_col + 1)}
    units_by_col   = {c: (ws.cell(ROW_UNIT, c).value or '') for c in range(2, max_col + 1)}

    # Detect table-definition rows and data start
    # A row is a table-definition row if it has at least one label in cols 2+
    # (col A gives the table/file name; empty col A → auto-named "table1", "table2"…)
    # Stop when col A == "Austria" (first data row).
    table_row_nums: list[int] = []
    data_start: int | None = None
    for r in range(ROW_LATEX, ws.max_row + 1):
        col_a = ws.cell(r, 1).value
        col_a_s = str(col_a).strip() if col_a is not None else ''
        if col_a_s.lower() == 'austria':
            data_start = r
            break
        has_labels = any(
            ws.cell(r, c).value and str(ws.cell(r, c).value).strip()
            for c in range(2, max_col + 1)
        )
        if has_labels:
            table_row_nums.append(r)

    if data_start is None:
        raise ValueError("'Austria' row not found — cannot locate data section.")
    if not table_row_nums:
        raise ValueError("No table-definition rows found between the header and 'Austria'.")

    # Read all data rows once
    all_data: list[tuple[str, dict]] = []
    for r in range(data_start, ws.max_row + 1):
        country = ws.cell(r, 1).value
        if country is None:
            continue
        row_vals = {c: ws.cell(r, c).value for c in range(2, max_col + 1)}
        all_data.append((str(country), row_vals))

    # Build per-table info
    tables: list[dict] = []
    for row_num in table_row_nums:
        latex_row = [ws.cell(row_num, c).value for c in range(1, max_col + 1)]
        auto_idx   = table_row_nums.index(row_num) + 1
        table_name = str(latex_row[0]).strip() if latex_row[0] and str(latex_row[0]).strip() else f'table{auto_idx}'

        included = [
            (c, str(latex_row[c - 1]).strip())
            for c in range(2, max_col + 1)
            if latex_row[c - 1] and str(latex_row[c - 1]).strip()
        ]
        if not included:
            continue

        source_groups: OrderedDict[str, list] = OrderedDict()
        for col, label in included:
            src = str(sources_by_col.get(col) or '')
            source_groups.setdefault(src, []).append((col, label))

        alignments = {}
        for col, _ in included:
            vals = [rv.get(col) for _, rv in all_data]
            numeric = all(isinstance(v, (int, float)) for v in vals if v not in (None, ''))
            alignments[col] = 'r' if numeric else 'c'

        data = [(country, {c: rv.get(c) for c, _ in included}) for country, rv in all_data]

        tables.append({
            'name': table_name,
            'country_label': table_name,
            'included': included,
            'source_groups': source_groups,
            'units_by_col': units_by_col,
            'alignments': alignments,
            'data': data,
        })

    return tables


def _caption_macro(table_name: str) -> str:
    """Derive a LaTeX macro name from the table name (letters only)."""
    safe = ''.join(c for c in table_name if c.isalpha())
    return f'\\{safe}caption'


def generate_longtable(info: dict, caption: str, label: str,
                        landscape: bool = False,
                        fontsize: str = '',
                        tabcolsep: str = '',
                        extra_cols: int = 0) -> str:
    included      = info['included']
    country_label = info['country_label']
    data          = info['data']
    table_name    = info.get('name', 'table')
    macro         = _caption_macro(table_name)

    todo_labels = [f'To be completed {i+1}' for i in range(extra_cols)]
    rng = random.Random(42)
    source_cols = [col for col, _ in included] if included else []
    todo_data: dict[str, list] = {country: [] for country, _ in data}
    for _ in range(extra_cols):
        src_col = rng.choice(source_cols) if source_cols else None
        for country, row_vals in data:
            val = row_vals.get(src_col) if src_col is not None else ''
            todo_data[country].append(val)

    total_cols = 1 + len(included) + extra_cols
    data_cols  = total_cols - 1
    col_spec = (
        r'>{\raggedright\arraybackslash}p{\firstcolw}'
        + (r' >{\centering\arraybackslash}p{\colw}' * data_cols)
    )

    pkg_comment = '% Requires: \\usepackage{longtable, booktabs, array}'
    if landscape:
        pkg_comment += ', pdflscape'

    out = []
    out.append(pkg_comment)
    if landscape:
        out.append('\\begin{landscape}')
    if fontsize:
        out.append(f'{{{fontsize}')
    if tabcolsep:
        out.append(f'\\setlength{{\\tabcolsep}}{{{tabcolsep}}}')
    # firstcolw fixed to fit "Luxembourg" on one line; colw splits the remainder
    # \ifdefined guards prevent "already defined" errors when multiple tables are included
    out.append(r'\ifdefined\firstcolw\else\newlength{\firstcolw}\fi')
    out.append(r'\setlength{\firstcolw}{2.2cm}')
    out.append(r'\ifdefined\colw\else\newlength{\colw}\fi')
    out.append(f'\\setlength{{\\colw}}{{\\dimexpr(\\linewidth - \\firstcolw)/{data_cols} - 2\\tabcolsep\\relax}}')

    # Caption macro: \providecommand so the parent doc can override with \renewcommand
    out.append(f'\\providecommand{{{macro}}}{{{escape_tex(caption)}}}')

    out.append(f'\\begin{{longtable}}{{{col_spec}}}')
    out.append(f'\\caption{{{macro}}}\\label{{{label}}} \\\\')

    out.append(r'\toprule')
    name_parts = (
        [escape_tex(country_label)]
        + [escape_tex(lbl) for _, lbl in included]
        + [f'\\textit{{{escape_tex(lbl)}}}' for lbl in todo_labels]
    )
    out.append(' & '.join(name_parts) + r' \\')
    out.append(r'\midrule')
    out.append(r'\endhead')

    out.append(r'\bottomrule')
    out.append(r'\endlastfoot')

    prev_was_total = False
    for country, row_vals in data:
        is_total = country.strip().lower() in TOTAL_KEYS
        if is_total:
            data_vals = [row_vals.get(col) for col, _ in included]
            if all(v is None or v == '' for v in data_vals):
                continue
        if is_total and not prev_was_total:
            out.append(r'\midrule')
        cells = (
            [escape_tex(country)]
            + [fmt_num(row_vals.get(col)) for col, _ in included]
            + [fmt_num(v) for v in todo_data.get(country, [0.0] * extra_cols)]
        )
        row_str = ' & '.join(cells) + r' \\'
        if is_total:
            row_str = ' & '.join(f'\\textbf{{{c}}}' for c in cells) + r' \\'
        out.append(row_str)
        prev_was_total = is_total

    out.append(r'\end{longtable}')
    if fontsize:
        out.append('}')
    if landscape:
        out.append('\\end{landscape}')
    return '\n'.join(out)


def main():
    ap = argparse.ArgumentParser(description='xlsx → LaTeX longtable(s)')
    ap.add_argument('xlsx', help='Input xlsx file')
    ap.add_argument('-o', '--output', help='Output path: single file (one table) or sibling dir (multiple tables)')
    ap.add_argument('--stdout', action='store_true', help='Print to stdout instead of file(s)')
    ap.add_argument('--caption', default='', help='Table caption (default: table name)')
    ap.add_argument('--label',   default='',  help='LaTeX label (default: tab:<name>)')
    ap.add_argument('--landscape', action='store_true', help='Wrap in landscape environment')
    ap.add_argument('--fontsize',  default='', help=r'Font size, e.g. \footnotesize')
    ap.add_argument('--tabcolsep', default='', help='Column padding, e.g. 3pt')
    ap.add_argument('--extra-cols', type=int, default=0, metavar='N',
                    help='Add N placeholder columns')
    args = ap.parse_args()

    tables = read_xlsx(Path(args.xlsx))

    gen_kwargs = dict(
        landscape=args.landscape,
        fontsize=args.fontsize,
        tabcolsep=args.tabcolsep,
        extra_cols=args.extra_cols,
    )

    if args.stdout or not args.output:
        for info in tables:
            caption = args.caption or info['name']
            label   = args.label   or f'tab:{info["name"]}'
            if len(tables) > 1:
                print(f'% === TABLE: {info["name"]} ===')
            print(generate_longtable(info, caption, label, **gen_kwargs))
        return

    out_path = Path(args.output)

    if len(tables) == 1:
        # Single table: write to -o directly (backwards compatible)
        info    = tables[0]
        caption = args.caption or info['name']
        label   = args.label   or f'tab:{info["name"]}'
        out_path.write_text(generate_longtable(info, caption, label, **gen_kwargs), encoding='utf-8')
        print(f'Saved: {out_path}')
    else:
        # Multiple tables: write <name>.tex into the directory of -o
        out_dir = out_path.parent if out_path.suffix == '.tex' else out_path
        out_dir.mkdir(parents=True, exist_ok=True)
        for info in tables:
            caption  = args.caption or info['name']
            label    = args.label   or f'tab:{info["name"]}'
            dest     = out_dir / f'{info["name"]}.tex'
            dest.write_text(generate_longtable(info, caption, label, **gen_kwargs), encoding='utf-8')
            print(f'Saved: {dest}')


if __name__ == '__main__':
    raise SystemExit(main())
